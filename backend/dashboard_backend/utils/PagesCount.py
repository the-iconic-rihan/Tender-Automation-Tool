
from pymongo import MongoClient
import os
import fitz
from ..logger import *
from openpyxl import load_workbook



class PagesCount():

    def __init__(self):
        pass

    def mongodb_connection(self):
        try:
            mongodb_uri = os.environ.get('MONGODB_URI')
            if not mongodb_uri:
                raise ValueError("MONGODB_URI is not defined in the environment.")

            client = MongoClient(mongodb_uri)
            db = client["Metadata"]
            collection = db["tender_metadata"]
            tender_page_count = db['tender_count_metadata']

            return client, collection,tender_page_count

        except Exception as e:
            exception_logger.error(f"Failed to connect to MongoDB: {str(e)}")
            raise

    def initialize_counts(self, existing_document):
        num_pages = existing_document.get("num_pages",0)
        sheet_count = existing_document.get("sheet_count",0)
        total_count = existing_document.get("total_count",0)
        return num_pages, sheet_count, total_count

    def create_default_document(self, tender_page_count, division):
        try:
            # Create and insert a new document with default values
            default_document = {
                'division': division,
                'num_pages': 0,
                'sheet_count': 0,
                'total_count': 0
            }
            tender_page_count.insert_one(default_document)
            audit_logger.info(f"New document created for division: {division}")

        except Exception as e:
            exception_logger.error(f"Error creating default document: {str(e)}")
            raise e


    def extract_number_of_pages(self, file, division, tender_number) :
        
        try:
            _, _, tender_page_count = self.mongodb_connection()
            # Find a document for the given division
            existing_document = tender_page_count.find_one({'division': division})
            if not existing_document:
                audit_logger.info(f"Creating a new document for division: {division}")
                self.create_default_document(tender_page_count, division)

            # Retrieve counts from the document
            existing_document = tender_page_count.find_one({'division': division})
            existing_document_data = {
                "num_pages" : existing_document.get("num_pages"),
                "sheet_count" : existing_document.get("sheet_count"),
                "total_count" : existing_document.get("total_count")

            }

            file_extension = os.path.splitext(file)[-1][1:]

            if file_extension.lower() == "pdf":
                with fitz.open(file) as pdf_document:
                    existing_document_data["num_pages"]+=pdf_document.page_count
            elif file_extension.lower() == "xlsx":
                workbook = load_workbook(file)
                existing_document_data["sheet_count"]+=len(workbook.sheetnames)
            else:
                exception_logger.error(f"Unsupported file type: {file_extension}")
                raise ValueError(f"Unsupported file type: {file_extension}")

            existing_document_data["total_count"] = existing_document_data["sheet_count"] + existing_document_data["num_pages"]

            filter_condition = {"division": division}
            update_data = {"$set": {"num_pages": existing_document_data["num_pages"], "sheet_count": existing_document_data["sheet_count"], "total_count": existing_document_data["total_count"]}}

            tender_page_count.update_one(filter_condition, update_data)

            audit_logger.info(f"Page counts updated for division: {division}, tender_number: {tender_number}")
            return existing_document_data["total_count"]

        except Exception as e:
            exception_logger.error(f"Error extracting page count: {str(e)}")
            raise e
    
    
    
    def extract_number_of_pages_per_tender(self, tender_number, file):
        try:
            audit_logger.info(f"file ===> {file}")
            _, collection,_ = self.mongodb_connection()

            existing_document = collection.find_one({"tender_number": tender_number})
            

            num_pages, sheet_count, total_count = self.initialize_counts(existing_document)

            file_extension = os.path.splitext(file)[-1][1:]

            if file_extension.lower() == "pdf":
                with fitz.open(file) as pdf_document:
                    num_pages += pdf_document.page_count
            elif file_extension.lower() == "xlsx":
                workbook = load_workbook(file)
                sheet_count += len(workbook.sheetnames)
            else:
                exception_logger.error(f"Unsupported file type: {file_extension}")
                raise ValueError(f"Unsupported file type: {file_extension}")
            
            
            total_count = sheet_count + num_pages

            filter_condition = {"tender_number": tender_number}
            update_data = {"$set": {"num_pages": num_pages, "sheet_count": sheet_count, "total_count": total_count}}

            collection.update_one(filter_condition, update_data, upsert=True)

            audit_logger.info(f"Total page count for {tender_number}: {total_count}")

            return total_count

        except Exception as e:
            exception_logger.error(f"Error extracting page count: {str(e)}")
            raise     
